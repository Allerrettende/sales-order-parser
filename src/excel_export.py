import pandas as pd

def orders_to_dataframe(orders_data):
    """
    将订单数据列表转换为 DataFrame
    """
    if not orders_data:
        return pd.DataFrame()
    
    all_rows = []
    
    for order_data in orders_data:
        header = order_data['header']
        customer = order_data['customer']
        items = order_data['items']
        
        # 为每个订单项创建一行数据
        for item in items:
            row = {
                # 表头信息
                "sales_order_no": header.get('sales_order_no', ''),
                "document_date": header.get('document_date', ''),
                "customer_no": header.get('customer_no', ''),
                "agent": header.get('agent', ''),
                "reference_quote_no": header.get('reference_quote_no', ''),
                "currency": header.get('currency', ''),
                "subtotal_amount": header.get('subtotal_amount', 0),
                
                # 客户信息
                "customer_name": customer.get('customer_name', ''),
                "customer_address": customer.get('customer_address', ''),
                "customer_postcode": customer.get('customer_postcode', ''),
                "customer_city": customer.get('customer_city', ''),
                "customer_country": customer.get('customer_country', ''),
                
                # 订单项信息
                "pos_number": item.get('pos_number', ''),
                "item_description": item.get('item_description', ''),
                "item_details": '\n'.join(item.get('item_details', [])) if item.get('item_details') else '',
                "quantity": item.get('quantity', 0),
                "unit": item.get('unit', ''),
                "unit_price": item.get('unit_price', 0),
                "amount": item.get('amount', 0),
                "discount":item.get('discount',0),
                "tax_code": item.get('tax_code', ''),
            }
            all_rows.append(row)
    
    # 创建 DataFrame
    df = pd.DataFrame(all_rows)
    
    # 定义列的显示顺序
    column_order = [
        "sales_order_no",
        "document_date",
        "customer_no",
        "customer_name",
        "customer_address",
        "customer_postcode",
        "customer_city",
        "customer_country",
        "agent",
        "reference_quote_no",
        "currency",
        "subtotal_amount",
        "pos_number",
        "item_description",
        "item_details",
        "quantity",
        "unit",
        "unit_price",
        "amount",
        "discount",
        "tax_code",
    ]
    
    # 只保留存在的列并按指定顺序排列
    existing_columns = [col for col in column_order if col in df.columns]
    df = df[existing_columns]
    
    return df

def apply_excel_formatting(worksheet, df, df_columns):
    """
    对 Excel 工作表应用格式化样式
    
    Args:
        worksheet: openpyxl 工作表对象
        df: 原始 DataFrame
        df_columns: DataFrame 的列名列表
    """
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    
    # 设置列宽 - 使用列名定义
    column_widths = {
        'sales_order_no': 18,
        'document_date': 15,
        'customer_no': 15,
        'customer_name': 30,
        'customer_address': 35,
        'customer_postcode': 15,
        'customer_city': 20,
        'customer_country': 15,
        'agent': 18,
        'reference_quote_no': 20,
        'currency': 12,
        'subtotal_amount': 18,
        'pos_number': 12,
        'item_description': 40,
        'item_details': 48,
        'quantity': 12,
        'unit': 12,
        'unit_price': 15,
        'amount': 15,
        'discount':15,
        'tax_code': 12,
    }
    
    # 为每一列设置宽度
    for col_idx, col_name in enumerate(df_columns, start=1):
        col_letter = get_column_letter(col_idx)
        if col_name in column_widths:
            worksheet.column_dimensions[col_letter].width = column_widths[col_name]
        else:
            # 如果没有指定宽度，根据内容自动调整（但限制最大宽度）
            max_length = max(
                df[col_name].astype(str).map(len).max(),
                len(col_name)
            )
            worksheet.column_dimensions[col_letter].width = min(max_length + 2, 50)
    
    # 设置表头样式
    header_font = Font(name='Arial', size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    for cell in worksheet[1]:  # 第一行是表头
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 定义需要右对齐和数字格式的列
    right_align_columns = ['subtotal_amount', 'quantity', 'unit_price', 'amount','discount']
    number_format_columns = ['subtotal_amount', 'unit_price', 'amount','discount']
    
    # 设置数据格式
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            # 默认左对齐
            cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # 获取列名
            col_name = df_columns[cell.column - 1]
            
            # 特定列右对齐
            if col_name in right_align_columns:
                cell.alignment = Alignment(horizontal='right', vertical='center')
            
            # 金额列设置数字格式
            if col_name in number_format_columns:
                cell.number_format = '#,##0.00'
    
    # 冻结首行
    worksheet.freeze_panes = 'A2'


def export_to_excel_with_formatting(df, output_path):
    """
    将 DataFrame 导出为格式化的 Excel 文件
    """
    if df.empty:
        print("❌ No data to export!")
        return
    
    # 创建 Excel writer
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # 写入数据
        df.to_excel(writer, sheet_name='Orders', index=False)
        
        # 获取工作表并应用格式化
        worksheet = writer.sheets['Orders']
        df_columns = df.columns.tolist()
        apply_excel_formatting(worksheet, df, df_columns)