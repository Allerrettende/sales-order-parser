
import pandas as pd

def apply_excel_formatting(worksheet, df, df_columns):
    """
    对 Excel 工作表应用格式化样式
    
    Args:
        worksheet: openpyxl 工作表对象
        df: 原始 DataFrame
        df_columns: DataFrame 的列名列表
    """
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    
    # 设置列宽 - 使用列名定义
    column_widths = {
        'sales_order_no': 12,
        'document_date': 12,
        'customer_no': 12,
        'customer_name': 30,
        'customer_address': 35,
        'customer_postcode': 8,
        'customer_city': 12,
        'customer_country': 8,
        'agent': 12,
        'reference_quote_no': 12,
        'currency': 8,
        'subtotal_amount': 12,
        'pos_number': 8,
        'item_description': 40,
        'item_details': 30,
        'quantity': 8,
        'unit': 8,
        'unit_price': 12,
        'amount': 12,
        'discount':12,
        'tax_code': 8,
    }
    
    # 为每一列设置宽度
    for col_idx, col_name in enumerate(df_columns, start=1):
        col_letter = get_column_letter(col_idx)
        if col_name in column_widths:
            worksheet.column_dimensions[col_letter].width = column_widths[col_name]
        else:
            # 如果没有指定宽度，根据内容自动调整（但限制最大宽度）
            max_length = max(
                df[col_name].astype(str).map(len).max(),
                len(col_name)
            )
            worksheet.column_dimensions[col_letter].width = min(max_length + 2, 50)
    
    # 设置表头样式
    header_font = Font(name='Arial', size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    for cell in worksheet[1]:  # 第一行是表头
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 定义需要右对齐和数字格式的列
    right_align_columns = ['subtotal_amount', 'quantity', 'unit_price', 'amount','discount']
    number_format_columns = ['subtotal_amount', 'unit_price', 'amount','discount']
    
    # 设置数据格式
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            # 默认左对齐
            cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # 获取列名
            col_name = df_columns[cell.column - 1]
            
            # 特定列右对齐
            if col_name in right_align_columns:
                cell.alignment = Alignment(horizontal='right', vertical='center')
            
            # 金额列设置数字格式
            if col_name in number_format_columns:
                cell.number_format = '#,##0.00'
    
    # 冻结首行
    worksheet.freeze_panes = 'D2'


def export_to_excel_with_formatting(df, output_path):
    """
    将 DataFrame 导出为格式化的 Excel 文件
    """
    if df.empty:
        print("❌ No data to export!")
        return
    
    # 创建 Excel writer
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # 写入数据
        df.to_excel(writer, sheet_name='Orders', index=False)
        
        # 获取工作表并应用格式化
        worksheet = writer.sheets['Orders']
        df_columns = df.columns.tolist()
        apply_excel_formatting(worksheet, df, df_columns)

# append multi excel-format orders to current list.
def append_multiple_orders(df, file_path, sheet_name='Orders', 
                          order_column='sales_order_no'):
    """
    批量检查并追加多个订单
    
    Args:
        df: 包含多个订单的数据
        file_path: Excel 文件路径
        sheet_name: 工作表名称
        order_column: 订单号列名
    """
    if df.empty:
        print("❌ No data to append!")
        return False
    
    try:
        # 获取所有新订单号
        new_orders = df[order_column].unique()
        print(f"📋 Preparing to add {len(new_orders)} orders")
        # print(f"📊 Total items: {len(df)}")
        
        # 检查文件是否存在
        try:
            existing_df = pd.read_excel(file_path, sheet_name=sheet_name)
            existing_orders = set(existing_df[order_column].unique())
            
            # 检查哪些订单已存在
            new_orders_set = set(new_orders)
            duplicate_orders = new_orders_set & existing_orders
            new_orders_to_add = new_orders_set - existing_orders
            
            if duplicate_orders:
                print(f"⚠️ Found {len(duplicate_orders)} orders already exist:")
                # ===========Debug======================
                # for order in list(duplicate_orders)[:5]:
                #     print(f"   - {order}")
                # if len(duplicate_orders) > 5:
                #     print(f"   ... and {len(duplicate_orders) - 5} more")
            
            if new_orders_to_add:
                print(f"✅ {len(new_orders_to_add)} new orders to add:")
                # for order in list(new_orders_to_add)[:5]:
                #     print(f"   - {order}")
                # if len(new_orders_to_add) > 5:
                #     print(f"   ... and {len(new_orders_to_add) - 5} more")
                
                # 只保留新订单的数据
                df_to_add = df[df[order_column].isin(new_orders_to_add)]
                
                # 追加数据
                combined_df = pd.concat([existing_df, df_to_add], ignore_index=True)
                
                with pd.ExcelWriter(file_path, engine='openpyxl', mode='w') as writer:
                    combined_df.to_excel(writer, sheet_name=sheet_name, index=False)
                    worksheet = writer.sheets[sheet_name]
                    df_columns = combined_df.columns.tolist()
                    apply_excel_formatting(worksheet, combined_df, df_columns)
                
                print(f"✅ Successfully added {len(new_orders_to_add)} new orders")
                print(f"📊 Total orders: {len(combined_df[order_column].unique())}")
                print(f"📊 Total rows: {len(combined_df)}")
                return True
            else:
                print("ℹ️ No new orders to add")
                return False
                
        except FileNotFoundError:
            # 文件不存在，创建新文件
            export_to_excel_with_formatting(df, file_path)
            print(f"✅ Created new file with {len(new_orders)} orders")
            return True
            
    except Exception as e:
        print(f"❌ Error: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

# Append data to master excel file from temp excel file with same structure.    

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import gc

def append_records_from_another_file(
    main_file,      # A文件：主表（已格式化）
    temp_file,      # B文件：中间表（新订单）
    sheet_name='Orders',
    order_column='sales_order_no'
):
    """
    从中间文件(B)追加新订单到主文件(A:Table)
    
    Args:
        main_file: 主文件路径（已格式化的正式表）
        temp_file: 中间文件路径（每日新订单列表）
        sheet_name: 工作表名称
        order_column: 订单号列名
    """
    
    try:
        # 1. 读取中间文件(B)的数据
        new_df = pd.read_excel(temp_file, sheet_name=sheet_name)
        if new_df.empty:
            return False
        
        # 2. 获取新订单的订单号列表
        new_orders = new_df[order_column].unique()
        
        # 3. 读取主文件(A)的现有订单号（只读，不加载全部数据）
        # print(f"📂 Reading main file: {main_file}")
        
        # 使用只读模式获取现有订单号
        from openpyxl import load_workbook
        
        # 方法1：使用pandas读取（简单，但会加载全部数据到内存）
        existing_df = pd.read_excel(main_file, sheet_name=sheet_name)
        existing_orders = set(existing_df[order_column].unique())
        
        # 方法2：使用openpyxl只读模式（内存友好）
        # existing_orders = get_existing_order_numbers(main_file, sheet_name, order_column)
        
        # 4. 检查哪些订单是新的
        new_orders_set = set(new_orders)
        duplicate_orders = new_orders_set & existing_orders # 交集
        new_orders_to_add = new_orders_set - existing_orders # 去集
        
        if duplicate_orders:
            print(f"⚠️ Found {len(duplicate_orders)} orders already exist:")
        
        if not new_orders_to_add:
            print("ℹ️ No new orders to add")
            return True
        
        print(f"✅ Found {len(new_orders_to_add)} new orders to add:")

        
        # 5. 筛选出新订单的数据
        df_to_add = new_df[new_df[order_column].isin(new_orders_to_add)]
        
        print(f"📊 Will add {len(df_to_add)} rows")
        
        # 6. 追加到主文件（保持格式）
        success = append_data_preserve_format(
            main_file, 
            df_to_add, 
            sheet_name, 
            order_column,
            existing_orders
        )
        
        if success:
            print("=" * 60)
            print("✅ Import completed successfully!")
            print(f"   Added {len(new_orders_to_add)} new orders")
            print(f"   Added {len(df_to_add)} rows")
            print("=" * 60)
        else:
            print("❌ Import failed")
        
        return success
        
    except FileNotFoundError as e:
        print(f"❌ File not found: {e}")
        return False
    except Exception as e:
        print(f"❌ Error: {str(e)}")
        import traceback
        traceback.print_exc()
        return False


def get_existing_order_numbers(file_path, sheet_name, order_column):
    """
    使用openpyxl只读模式获取现有订单号（内存友好）
    """
    from openpyxl import load_workbook
    
    existing_orders = set()
    
    try:
        # 使用只读模式，不加载全部数据
        book = load_workbook(file_path, read_only=True, data_only=True)
        
        if sheet_name not in book.sheetnames:
            book.close()
            return set()
        
        sheet = book[sheet_name]
        
        # 找到订单号列的索引
        headers = []
        for col_idx, cell in enumerate(sheet[1], start=1):
            headers.append(cell.value)
        
        try:
            order_col_idx = headers.index(order_column) + 1
        except ValueError:
            print(f"⚠️ Column '{order_column}' not found in main file")
            book.close()
            return set()
        
        # 只读取订单号列
        for row in sheet.iter_rows(min_row=2, max_col=order_col_idx, values_only=True):
            if row and row[order_col_idx - 1] is not None:
                existing_orders.add(row[order_col_idx - 1])
        
        book.close()
        
    except Exception as e:
        print(f"⚠️ Could not read main file: {e}")
        return set()
    
    return existing_orders


def append_data_preserve_format(file_path, df_to_add, sheet_name, order_column, existing_orders):
    """
    追加数据到主文件，完全保留原有格式
    """
    from openpyxl import load_workbook
    from openpyxl.utils import range_boundaries
    
    try:
        print("📝 Appending data while preserving formatting...")
        
        # 加载工作簿（保留所有格式）
        book = load_workbook(file_path)
        
        if sheet_name not in book.sheetnames:
            print(f"❌ Sheet '{sheet_name}' not found in main file")
            book.close()
            return False
        
        sheet = book[sheet_name]
        
        # 获取表头
        headers = [cell.value for cell in sheet[1]]
        
        # 验证列匹配
        df_columns = df_to_add.columns.tolist()
        if len(headers) != len(df_columns):
            print(f"⚠️ Column count mismatch: Excel has {len(headers)}, DataFrame has {len(df_columns)}")
            # 尝试匹配列名
            matched_columns = []
            for h in headers:
                if h in df_columns:
                    matched_columns.append(h)
                else:
                    matched_columns.append(None)
            print(f"   Matched columns: {matched_columns}")
        else:
            matched_columns = headers
        
        # 获取当前最后一行
        last_row = sheet.max_row
       
        # 追加数据
        for row_idx, (_, row_data) in enumerate(df_to_add.iterrows(), start=last_row + 1):
            for col_idx, col_name in enumerate(matched_columns, start=1):
                if col_name is not None and col_name in row_data:
                    value = row_data[col_name]
                    cell = sheet.cell(row=row_idx, column=col_idx, value=value)

        # ===================== 如果当前excel sheet中的是Table.
        # 调整Table范围
        # 获取当前最后一行
        new_last_row = sheet.max_row   
        # 假设你的 Table 名称是 'Table1'
        table_name = 'Table1'
        if table_name in sheet.tables:
            table = sheet.tables[table_name]
            
            # 解析旧范围，保留起始行，只更新结束行
            # 例如，从旧范围 "A1:N2" 中解析出起始列 "A" 和结束列 "N"
            start_cell = table.ref.split(':')[0] 
            end_col = table.ref.split(':')[1] 
            # 从结束列引用中提取列字母，例如从 "N2" 提取 "N"
            # 更稳健的方式是用 openpyxl.utils 解析，但这里用简单字符串处理
            import re
            end_col_letter = re.sub(r'[\d]+', '', end_col) 
            
            # 构建新的范围字符串，例如 "A1:N101"
            new_ref = f"{start_cell}:{end_col_letter}{new_last_row}"
            table.ref = new_ref
            print(f"✅ Table range updated to: {new_ref}")
        else:
            print(f"⚠️ Table '{table_name}' not found.")

        # =====================调整Table结束。

      # 保存文件
        book.save(file_path)
        book.close()
        
        print(f"✅ Data appended successfully")
        return True
        
    except Exception as e:
        print(f"❌ Error appending data: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

